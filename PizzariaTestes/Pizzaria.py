from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.keys import Keys
import time
from openpyxl  import Workbook,load_workbook
import os
import urllib.request
from selenium.webdriver.support.select import Select


#PARAMETROS#
pedidos = []
#PARAMETROS#

#CALCULANDO PRIMEIRA LINHA
resultadoPedidos = 'ResultadoPedidos.xlsx'
wb = load_workbook(resultadoPedidos)
ws = wb.active
numComeco = len(ws['C'])+1
print ("Primeira linha: ", numComeco-1)
wb.save(resultadoPedidos)
#CALCULANDO PRIMEIRA LINHA#

#CALCULANDO ÚLTIMA LINHA#
entradas = "EntradasPedidos.xlsx"
wb = load_workbook(entradas)
ws = wb.active
numFinal = len(ws['A'])+1
print("Ultima linha: ", numFinal-2)
wb.save(entradas)
#CALCULANDO ÚLTIMA LINHA#

#ABRINDO ARQUIVOS#
arqPedidos = load_workbook(resultadoPedidos)
sheetPedidos = arqPedidos.active
arqEntradas = load_workbook(entradas)
sheetEntradas = arqEntradas.active
#ABRINDO ARQUIVOS#

#COPIANDO pedidos#
for linha in range(numComeco,numFinal):
	celPedidos = sheetEntradas.cell(row=linha, column=1)
	print("Pedidos carregados: ", celPedidos.value)
	pedidos = celPedidos.value

	celBebidas = sheetEntradas.cell(row=linha, column=2)
	bebidas = celBebidas.value

	celTamanhos = sheetEntradas.cell(row=linha, column=3)
	tamanhos = celTamanhos.value
	
	celMassas = sheetEntradas.cell(row=linha, column=4)
	massas = celMassas.value

	celBordas = sheetEntradas.cell(row=linha, column=5)
	bordas = celBordas.value

	celSabores = sheetEntradas.cell(row=linha, column=6)
	sabores = celSabores.value
#COPIANDO pedidos#

#ABRINDO NAVEGADOR#
driver = webdriver.Chrome('C:/chromedriver.exe')
print ("Abrindo navegador\n")
driver.get("http://localhost:4200/")
print ("Acessando Pizzaria...\n")
driver.maximize_window()
#ABRINDO NAVEGADOR#

#LOGIN#
def login():
	senhaRobo = readlog = open(os.getcwd()+"\\LoginPizzaria.txt").readlines()
	driver.execute_script("document.getElementById('user').setAttribute('value','"+readlog[0].strip()+"')")
	senhaRobo = driver.execute_script("document.getElementById('pass').value = '"+readlog[1]+"'")
	driver.implicitly_wait(10)
	logar = driver.find_element_by_xpath("""/html/body/app-root/app-hero-detail/div[2]/div[1]/button""").click()

login()
#LOGIN#


def criarPedido(bebida, tamanho, massa, borda, sabor):
	time.sleep(2)
	driver.find_element_by_xpath("""/html/body/app-root/app-home-page/header/nav/button""").click()
	time.sleep(2)

	selectBebida = driver.find_element_by_id('bebida')
	objectBebida = Select(selectBebida)
	objectBebida.select_by_visible_text(bebida)
	# tamanho
	selectTamanho = driver.find_element_by_id('tamanho')
	objectTamanho = Select(selectTamanho)
	objectTamanho.select_by_visible_text(tamanho)
	# massa
	selectMassa = driver.find_element_by_id('massa')
	objectMassa = Select(selectMassa)
	objectMassa.select_by_visible_text(massa)
	# borda
	selectBorda = driver.find_element_by_id('borda')
	objectBorda = Select(selectBorda)
	objectBorda.select_by_visible_text(borda)
	# sabor
	selectSabor = driver.find_element_by_id('sabor')
	objectSabor = Select(selectSabor)
	objectSabor.select_by_visible_text(sabor)

	# concluir pedido
	driver.find_element_by_xpath("""//*[@id="exampleModal"]/div/div/div[3]/button[2]""").click()
	
	

for linhaContrato in range(numComeco,numFinal):
	print(linhaContrato-1, '/', numFinal-2)
	celPedidos = sheetEntradas.cell(row=linhaContrato, column=1) 
	celBebidas = sheetEntradas.cell(row=linhaContrato, column=2) 
	celTamanhos = sheetEntradas.cell(row=linhaContrato, column=3) 
	celMassas = sheetEntradas.cell(row=linhaContrato, column=4) 
	celBordas = sheetEntradas.cell(row=linhaContrato, column=5) 
	celSabores = sheetEntradas.cell(row=linhaContrato, column=6) 
	criarPedido(celBebidas.value, celTamanhos.value, celMassas.value, celBordas.value, celSabores.value)

	sheetPedidos.cell(row=linhaContrato, column=1).value = celPedidos.value
	sheetPedidos.cell(row=linhaContrato, column=2).value = celBebidas.value
	sheetPedidos.cell(row=linhaContrato, column=3).value = celTamanhos.value
	sheetPedidos.cell(row=linhaContrato, column=4).value = celMassas.value
	sheetPedidos.cell(row=linhaContrato, column=5).value = celBordas.value
	sheetPedidos.cell(row=linhaContrato, column=6).value = celSabores.value

	arqPedidos.save(resultadoPedidos)

# ir para pedidos

time.sleep(2)
driver.find_element_by_xpath("""/html/body/app-root/app-home-page/div[1]/div/div/div[4]""").click()
print("Processo concluido com sucesso!")
